#!/usr/bin/env python3
"""
Paso 3.2 - Analisis del benchmark de cache-aside con Redis.

Lee resultados_benchmark.csv (generado por benchmark.sh) y calcula:
  - Media y desviacion estandar de tiempos sin cache y con cache
  - Percentil 95 (P95) de cada serie
  - Speedup S = Tsin / Tcon
  - Si S > 2 (umbral tipico que justifica la complejidad de Redis, segun
    el criterio de verificacion de la guia)

Genera:
  - resultados_benchmark.md  (tabla lista para pegar en el informe)
  - resultados_benchmark.xlsx (misma tabla en Excel, con formulas)

Uso:
    python3 analizar_benchmark.py [ruta_csv]
"""

import csv
import statistics
import sys
from pathlib import Path

CSV_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("resultados_benchmark.csv")


def percentil_95(valores):
    datos = sorted(valores)
    if len(datos) == 1:
        return datos[0]
    k = 0.95 * (len(datos) - 1)
    f = int(k)
    c = min(f + 1, len(datos) - 1)
    if f == c:
        return datos[f]
    return datos[f] + (datos[c] - datos[f]) * (k - f)


def main():
    if not CSV_PATH.exists():
        print(f"ERROR: no se encontro {CSV_PATH}. Ejecuta primero benchmark.sh")
        sys.exit(1)

    sin_cache, con_cache = [], []
    with open(CSV_PATH, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sin_cache.append(float(row["sin_cache_ms"]))
            con_cache.append(float(row["con_cache_ms"]))

    n = len(sin_cache)
    media_sin = statistics.mean(sin_cache)
    media_con = statistics.mean(con_cache)
    desv_sin = statistics.stdev(sin_cache) if n > 1 else 0.0
    desv_con = statistics.stdev(con_cache) if n > 1 else 0.0
    p95_sin = percentil_95(sin_cache)
    p95_con = percentil_95(con_cache)
    speedup = media_sin / media_con if media_con > 0 else float("inf")

    print("=" * 60)
    print(f"BENCHMARK CACHE-ASIDE REDIS (n={n} repeticiones)")
    print("=" * 60)
    print(f"{'Metrica':<25}{'Sin cache (ms)':>17}{'Con cache (ms)':>17}")
    print(f"{'Media':<25}{media_sin:>17.2f}{media_con:>17.2f}")
    print(f"{'Desv. estandar':<25}{desv_sin:>17.2f}{desv_con:>17.2f}")
    print(f"{'P95':<25}{p95_sin:>17.2f}{p95_con:>17.2f}")
    print("-" * 60)
    print(f"Speedup S = Tsin/Tcon = {speedup:.2f}x")
    umbral = "SI" if speedup > 2 else "NO"
    print(f"Justifica complejidad de Redis (S > 2)?: {umbral}")
    print("=" * 60)

    # --- Markdown para el informe ---
    md_lines = [
        "## Resultados del Benchmark de Cache (Paso 3.2)\n",
        f"n = {n} repeticiones sobre `GET /api/entidades?page=0&size=20&sort=precio,desc`\n",
        "| Iteracion | Sin cache (ms) | Con cache (ms) |",
        "|---|---|---|",
    ]
    for i, (s, c) in enumerate(zip(sin_cache, con_cache), start=1):
        md_lines.append(f"| {i} | {s:.2f} | {c:.2f} |")
    md_lines += [
        f"| **Media** | **{media_sin:.2f}** | **{media_con:.2f}** |",
        f"| **Desv. estandar** | {desv_sin:.2f} | {desv_con:.2f} |",
        f"| **P95** | {p95_sin:.2f} | {p95_con:.2f} |",
        "",
        f"**Speedup S = Tsin / Tcon = {speedup:.2f}x**\n",
        f"¿Justifica la complejidad operacional de Redis (umbral S > 2)? **{umbral}**\n",
    ]
    Path("resultados_benchmark.md").write_text("\n".join(md_lines), encoding="utf-8")
    print("\nGenerado: resultados_benchmark.md")

    # --- Excel con formulas (para quien prefiera entregar tabla en xlsx) ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Benchmark Cache"

        headers = ["Iteracion", "Sin cache (ms)", "Con cache (ms)"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, name="Arial")
            cell.fill = PatternFill("solid", fgColor="1F6FEB")
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")

        for i, (s, c) in enumerate(zip(sin_cache, con_cache), start=2):
            ws.cell(row=i, column=1, value=i - 1).font = Font(name="Arial")
            ws.cell(row=i, column=2, value=s).font = Font(name="Arial")
            ws.cell(row=i, column=3, value=c).font = Font(name="Arial")

        fila_media = n + 2
        ws.cell(row=fila_media, column=1, value="Media").font = Font(bold=True, name="Arial")
        ws.cell(row=fila_media, column=2, value=f"=AVERAGE(B2:B{n+1})").font = Font(name="Arial")
        ws.cell(row=fila_media, column=3, value=f"=AVERAGE(C2:C{n+1})").font = Font(name="Arial")

        fila_desv = n + 3
        ws.cell(row=fila_desv, column=1, value="Desv. estandar").font = Font(bold=True, name="Arial")
        ws.cell(row=fila_desv, column=2, value=f"=STDEV(B2:B{n+1})").font = Font(name="Arial")
        ws.cell(row=fila_desv, column=3, value=f"=STDEV(C2:C{n+1})").font = Font(name="Arial")

        fila_speedup = n + 5
        ws.cell(row=fila_speedup, column=1, value="Speedup S = Tsin/Tcon").font = Font(bold=True, name="Arial")
        ws.cell(row=fila_speedup, column=2, value=f"=B{fila_media}/C{fila_media}").font = Font(bold=True, name="Arial")

        for col in ("A", "B", "C"):
            ws.column_dimensions[col].width = 20

        wb.save("resultados_benchmark.xlsx")
        print("Generado: resultados_benchmark.xlsx (recuerda ejecutar recalc.py si usas el skill de xlsx)")
    except ImportError:
        print("openpyxl no instalado; se omitio la generacion de .xlsx (el .md es suficiente).")


if __name__ == "__main__":
    main()
